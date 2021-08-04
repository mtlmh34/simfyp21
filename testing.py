import bs4
import requests
import openpyxl
import re
# Import Module
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from functions import mainFunctions

import xlsxwriter

from tqdm import tqdm

for x in tqdm(range(100)):
    print(x)

'''
def get_blacklist_endpoints():

    data = []
    blacklist_endpoints = get_blacklist_endpoints_request().get('SearchResult', {})

    if blacklist_endpoints.get('total', 0) < 1:
        demisto.results('No endpoints were found.')

    endpoints = blacklist_endpoints.get('resources', [])

    for endpoint in endpoints:
        data.append({
            'ID': endpoint.get('id'),
            'Name': endpoint.get('name'),
            'GroupName': 'Blacklist'
        })

    context = {
        'CiscoISE.Endpoint(val.ID && val.ID === obj.ID)': data
    }

    return_outputs(tableToMarkdown('CiscoISE Blacklist Endpoints', data, removeNull=True), context, endpoints)


get_blacklist_endpoints()
print(listing)
print("===================================================")
'''
'''
if not os.path.isfile('blacklist.xlsx'):
    excel = xlsxwriter.Workbook('blacklist.xlsx')
    worksheet = workbook.add_worksheet()
    workbook.close()

username = "tekkwang1996@gmail.com"
from openpyxl import load_workbook

wb = load_workbook("logs.xlsx")  # open an Excel file and return a workbook

workbook = xlsxwriter.Workbook('Expenses01.xlsx')
worksheet = workbook.add_worksheet(username)
workbook.close()

if username in wb.sheetnames:
    print(username + ' exists')
else:
    print(username + ' does not exists')
    wb.create_sheet(username)
    print(username + ' created')

ws = wb[username]
wb.save("logs.xlsx")
'''

'''
from openpyxl import load_workbook
wb2 = load_workbook('template.xlsx')
wb2.create_sheet('sid1')
wb2.save('template.xlsx')

    
    
# HTML Document
HTML_DOC = """
            """


# Function to remove tags
def remove_tags(html):
    # parse html content
    soup = BeautifulSoup(html, "html.parser")

    for data in soup(['style', 'script']):
        # Remove tags
        data.decompose()

    # return data by retrieving the tag content
    return ' '.join(soup.stripped_strings)


# Retrieve data from excel
workbook = load_workbook("logs.xlsx")
source = workbook["Sheet1"]
count = 1

myworkbook = openpyxl.load_workbook('logs.xlsx')
worksheet = myworkbook.active

for cell in source['C']:
    HTML_DOC = cell.value
    if not HTML_DOC:
        print("Empty")
    else:
        display = remove_tags(HTML_DOC)
        place = 'G' + str(count)
        mystring = display.replace("_x000D_", " ")
        worksheet[place] = mystring

    count += 1

myworkbook.save('logs.xlsx')
myworkbook.close()
workbook.close()
'''
'''
   @staticmethod
    def spelling_check():
        # Function to check if there is any spelling mistake #
        # import TextBlob
        from textblob import TextBlob
        from openpyxl import Workbook, load_workbook

        # Retrieve data from excel
        workbook = load_workbook("logs.xlsx")
        source = workbook["Sheet1"]

        # Print the contents
        for cell in source['A']:
            mispell_word = cell.value
            textBlb = TextBlob(mispell_word)
            textCorrect = textBlb.correct()
            if mispell_word != textCorrect:
                print(mispell_word, " || ", textCorrect)

    @staticmethod
    def email_valid():
        # Function to check if email address is legitimate
        # Import py3-validate-email package
        from validate_email import validate_email
        from openpyxl import load_workbook

        # Retrieve data from excel
        workbook = load_workbook("logs.xlsx")
        source = workbook["Sheet1"]
        # Print the contents
        for cell in source['D']:
            emailAd = cell.value
            # To only check string that contain @
            if "@" in emailAd:
                # To get only the email address
                emailAddress = emailAd.split('<')
                emailAd = emailAddress[1]
                emailAd = emailAd[:-1]
                print(emailAd)
                is_valid = validate_email(email_address=emailAd, smtp_timeout=10, dns_timeout=10)
                # is_valid = validate_email(email_address=emailAd, check_regex=True, check_mx=True,
                #                           from_address='my@from.addr.ess', helo_host='my.host.name',
                #                           smtp_timeout=10, dns_timeout=10, use_blacklist=True)
                
                print(is_valid)
                print("==============")
'''
