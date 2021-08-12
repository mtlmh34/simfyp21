import os
import re

from flask import Flask, render_template, url_for, request, flash, redirect

import imaplib
# For connection
import easyimap as e
import smtplib
import logging
import xlsxwriter
from openpyxl import load_workbook
# For Machine Learning
import pandas as pd
from joblib import dump, load
import numpy as np

import nltk

nltk.download('punkt')
nltk.download('stopwords')
from nltk import PorterStemmer, word_tokenize
from nltk.corpus import stopwords
from functions import mainFunctions
from mlpfeature import html_exists
from mlpfeature import count_domain
from mlpfeature import count_dots
from mlpfeature import account_exists
from mlpfeature import paypal_exists
from mlpfeature import login_exists
from mlpfeature import bank_exists

# Import function from another python file
from sklearn.feature_extraction.text import CountVectorizer

app = Flask(__name__)


@app.route('/', methods=['GET', 'POST'])
def login():
    loading_gif = url_for('static', filename='imgload.gif')
    image_file = url_for('static', filename='login.png')  # image
    # declare variables
    global server
    global username
    global password
    global imap_url
    global platform
    try:
        if request.method == 'POST':
            username = request.form['email']  # id='email' from html form
            password = request.form['password']
            platform = request.form['platform']

            # Authenticates and retrieves email
            if platform == 'Gmail':
                imap_url = 'imap.gmail.com'
            else:
                imap_url = 'outlook.office365.com'
            server = e.connect(imap_url, username, password)

            return redirect('/inbox')
        else:
            return render_template('index.html', image_file=image_file, loading_gif=loading_gif)

    except imaplib.IMAP4.error:
        error = "invalid credentials"
        return render_template('index.html', image_file=image_file, error=error)


##############################################################################################
#ML Data processing functions
def cleaning(string):
    string = re.sub("[^0-9a-zA-Z\ ]", "", str(string))
    string = string.lower()
    string = string.strip()

    return string



def stem(string):
    tokenized = word_tokenize(string)
    stemmed = []
    stemmer = PorterStemmer()

    for word in tokenized:
        stemmed.append(stemmer.stem(word))

    return ' '.join(stemmed)


def remove_stopwords(string):
    STOP_WORDS = set(stopwords.words('english'))

    tokenized = word_tokenize(string)
    filtered = []

    for word in tokenized:
        if word not in STOP_WORDS:
            filtered.append(word)

    return " ".join(filtered)


######################################################################################

@app.route('/inbox')
# retrieve and log emails to csv and perform ML prediction
def email():
    global email
    global subject_list
    global body_list
    global email_address_list
    global percentage_list
    global result_list
    global model  # ML Model id 0,1 or 2
    global model_string # name of model
    global spellingResult
    global emailValidResultList
    global attachmentResultList
    global linkResultList
    global htmlExistsList
    global domainCountList
    global dotCountList
    global accountExistsList
    global paypalExistsList
    global loginExistsList
    global bankExistsList
    global userBlacklist
    global userWhitelist


    # check if user have blacklist and whitelist
    wbB = load_workbook('blacklist.xlsx')
    if username in wbB.sheetnames:
        sheetB = wbB[username]
    else:
        # wb.create_sheet(username)
        wbB.copy_worksheet(wbB["blacklist"]).title = username
        wbB.save('blacklist.xlsx')
        sheetB = wbB[username]

    wbW = load_workbook('whitelist.xlsx')
    if username in wbW.sheetnames:
        sheetW = wbW[username]
    else:
        # wb.create_sheet(username)
        wbW.copy_worksheet(wbW["whitelist"]).title = username
        wbW.save('whitelist.xlsx')
        sheetW = wbW[username]

    userBlacklist = []
    row_countB = sheetB.max_row
    for i in range(1, row_countB + 1):
        for k in range(1, 3):
            if k == 1:
                emailadd = sheetB.cell(row=i, column=k).value
                userBlacklist.append(emailadd)
    userWhitelist = []
    row_countW = sheetW.max_row
    for i in range(1, row_countW + 1):
        for k in range(1, 3):
            if k == 1:
                emailadd = sheetW.cell(row=i, column=k).value
                userWhitelist.append(emailadd)


    # 0-naivebayes, 1-MLP, 2-randomforest
    wb = load_workbook('model.xlsx')
    sheet = wb["Sheet1"]  # values will be saved to excel sheet"blacklist"
    model = sheet['A1'].value
    print("MODEL should be a number  ", model)
    wb.close()

    if model == 0:
        model_string = "Naive Bayes"
    if model == 1:
        model_string = "Multilayer Perceptron (MLP)"
    if model == 2:
        model_string = "Random Forest"

    percentage_list = []
    subject_list = []
    body_list = []
    email_address_list = []
    result_list = []
    emailValidResultList = []
    attachmentResultList = []
    linkResultList = []
    htmlExistsList = []
    domainCountList = []
    dotCountList = []
    accountExistsList = []
    paypalExistsList = []
    loginExistsList = []
    bankExistsList = []

    # Authenticates and retrieves email

    # logging emails
    logger = logging.getLogger('logger1')
    handler = logging.FileHandler('log1.txt')
    logger.addHandler(handler)
    logger.setLevel(logging.INFO)

    server = e.connect(imap_url, username, password)
    # inbox = server.listup()
    inbox = server.listids()
    if os.path.isfile('logs.xlsx'):
        # Check if email sheet exist
        wb = load_workbook("logs.xlsx")  # open an Excel file and return a workbook
        # check if email sheet exist
        if username in wb.sheetnames:  # user have login before
            print(username + ' exists')
            # server = e.connect(imap_url, username, password)
            email = server.mail(server.listids()[0])
            ws = wb[username]
            excelRow = 2
            if ws["A2"].value == email.title and ws["B2"].value == email.from_addr: #user have no new email
                print("This is working")
                for x in range(1, ws.max_row):
                    # read from excel file
                    excelColumn = 'A'
                    excelPosition = excelColumn + str(excelRow)
                    subject_list.append(ws[excelPosition].value)
                    print(excelPosition)

                    excelColumn = chr(ord(excelColumn) + 1)
                    excelPosition = excelColumn + str(excelRow)
                    email_address_list.append(ws[excelPosition].value)
                    print(excelPosition)

                    excelColumn = chr(ord(excelColumn) + 1)
                    excelPosition = excelColumn + str(excelRow)
                    body_list.append(ws[excelPosition].value)
                    print(excelPosition)

                    excelColumn = chr(ord(excelColumn) + 1)
                    excelPosition = excelColumn + str(excelRow)
                    print(excelPosition)

                    excelColumn = chr(ord(excelColumn) + 1)
                    excelPosition = excelColumn + str(excelRow)
                    result_list.append(ws[excelPosition].value)
                    print(excelPosition)

                    excelColumn = chr(ord(excelColumn) + 1)
                    excelPosition = excelColumn + str(excelRow)
                    print(excelPosition)

                    excelColumn = chr(ord(excelColumn) + 1)
                    excelPosition = excelColumn + str(excelRow)
                    print(excelPosition)

                    excelColumn = chr(ord(excelColumn) + 1)
                    excelPosition = excelColumn + str(excelRow)
                    print(excelPosition)

                    excelColumn = chr(ord(excelColumn) + 1)
                    excelPosition = excelColumn + str(excelRow)
                    percentage_list.append(ws[excelPosition].value)
                    print(excelPosition)

                    excelColumn = chr(ord(excelColumn) + 1)
                    excelPosition = excelColumn + str(excelRow)
                    emailValidResultList.append(ws[excelPosition].value)
                    print(excelPosition)

                    excelColumn = chr(ord(excelColumn) + 1)
                    excelPosition = excelColumn + str(excelRow)
                    attachmentResultList.append(ws[excelPosition].value)
                    print(excelPosition)

                    excelColumn = chr(ord(excelColumn) + 1)
                    excelPosition = excelColumn + str(excelRow)
                    linkResultList.append(ws[excelPosition].value)
                    print(excelPosition)

                    excelColumn = chr(ord(excelColumn) + 1)
                    excelPosition = excelColumn + str(excelRow)
                    htmlExistsList.append(ws[excelPosition].value)
                    print(excelPosition)

                    excelColumn = chr(ord(excelColumn) + 1)
                    excelPosition = excelColumn + str(excelRow)
                    domainCountList.append(ws[excelPosition].value)
                    print(excelPosition)

                    excelColumn = chr(ord(excelColumn) + 1)
                    excelPosition = excelColumn + str(excelRow)
                    dotCountList.append(ws[excelPosition].value)
                    print(excelPosition)

                    excelColumn = chr(ord(excelColumn) + 1)
                    excelPosition = excelColumn + str(excelRow)
                    accountExistsList.append(ws[excelPosition].value)
                    print(excelPosition)

                    excelColumn = chr(ord(excelColumn) + 1)
                    excelPosition = excelColumn + str(excelRow)
                    paypalExistsList.append(ws[excelPosition].value)
                    print(excelPosition)

                    excelColumn = chr(ord(excelColumn) + 1)
                    excelPosition = excelColumn + str(excelRow)
                    loginExistsList.append(ws[excelPosition].value)
                    print(excelPosition)

                    excelColumn = chr(ord(excelColumn) + 1)
                    excelPosition = excelColumn + str(excelRow)
                    bankExistsList.append(ws[excelPosition].value)
                    print(excelPosition)

                    excelRow += 1
                    wb.save("logs.xlsx")

            else:  # user have new email
                print("new email found")
                '''checkTitle = ws["A2"].value
                checkAddr = ws["B2"].value
                # server = e.connect(imap_url, username, password)
                # inbox = server.listids()
                for x in range(0, len(inbox)):  # change 10 to len(inbox) to get 100 mails
                    email = server.mail(server.listids()[x])
                    if checkTitle == email.title and checkAddr == email.from_addr:
                        break
                    else:
                        # ML
                        string = body
                        string = cleaning(string)
                        string = stem(string)
                        string = remove_stopwords(string)

                        # dataFrame with only the email body that has been cleaned for use with Naive Bayes only
                        n_df = pd.DataFrame({'text': string}, index=[0])

                        # dataFrame for MLP and RandomForest model (Data not cleaned on purpose)
                        df = pd.DataFrame({'text': body}, index=[0])

                        # dataFrame with columns extracted from email
                        mlp_df = pd.DataFrame(columns=['HtmlExists', 'DomainCount', 'DotCount',
                                                       'AccountExists', 'PaypalExists', 'LoginExists', 'BankExists'])

                        mlp_df['HtmlExists'] = df['text'].apply(html_exists)
                        mlp_df['DomainCount'] = df['text'].apply(count_domain)
                        mlp_df['DotCount'] = df['text'].apply(count_dots)
                        mlp_df['AccountExists'] = df['text'].apply(account_exists)
                        mlp_df['PaypalExists'] = df['text'].apply(paypal_exists)
                        mlp_df['LoginExists'] = df['text'].apply(login_exists)
                        mlp_df['BankExists'] = df['text'].apply(bank_exists)
                        print(mlp_df)

                        # save columns to variable
                        htmlExists = mlp_df.iloc[0]['HtmlExists']
                        domainCount = mlp_df.iloc[0]['DomainCount']
                        dotCount = mlp_df.iloc[0]['DotCount']
                        accountExists = mlp_df.iloc[0]['AccountExists']
                        paypalExists = mlp_df.iloc[0]['PaypalExists']
                        loginExists = mlp_df.iloc[0]['LoginExists']
                        bankExists = mlp_df.iloc[0]['BankExists']

                        # append to list
                        htmlExistsList.append(htmlExists)
                        domainCountList.append(domainCount)
                        dotCountList.append(dotCount)
                        accountExistsList.append(accountExists)
                        paypalExistsList.append(paypalExists)
                        loginExistsList.append(loginExists)
                        bankExistsList.append(bankExists)

                        if model == 0:  # naive bayes
                            vectorizer = load(r'naivebayesVectorizer.joblib')  # load vectorizer
                            nbclf = load(r'naivebayes.joblib')  # load the naivebayes ml model
                            x_matrix = n_df['text']
                            x_matrix = vectorizer.transform(n_df['text'])
                            my_prediction = nbclf.predict(x_matrix)
                            percentage = nbclf.predict_proba(x_matrix)
                            print(my_prediction, "MY PREDICTION")

                        if model == 1:  # MLP
                            nbclf = load(r'mlp.joblib')
                            my_prediction = nbclf.predict(mlp_df)
                            percentage = nbclf.predict_proba(mlp_df)
                            print(my_prediction, "MY PREDICTION")

                        if model == 2:  # Random Forest
                            nbclf = load(r'randomforest.joblib')
                            my_prediction = nbclf.predict(mlp_df)
                            percentage = nbclf.predict_proba(mlp_df)
                            print(my_prediction, "MY PREDICTION")

                        np.set_printoptions(formatter={'float_kind': '{:f}'.format})

                        if my_prediction == 1:
                            ml_result = 'Phishing'
                            percentage = format(percentage[0][1], '.12f')  # to 12decimal place
                            percentage = float(percentage) * 100  # convert to percent
                            percentage = str(percentage) + '%'
                            percentage_list.append(percentage)
                        elif my_prediction == 0:
                            ml_result = 'Non-Phishing'
                            percentage = format(percentage[0][0], '.12f')  # to 12decimal place
                            percentage = float(percentage) * 100  # convert to percent
                            percentage = str(percentage) + '%'
                            percentage_list.append(percentage)

                        # logger.info("Email attachment: ")
                        # logger.info(email.attachments)
                        emailAttachment = email.attachments
                        if not emailAttachment:
                            emailAttach = "-"
                        else:
                            attachment = emailAttachment[0]
                            attachment = str(attachment)
                            attach = attachment.split(',')
                            emailAttach = str(attach[0])
                            emailAttach = emailAttach[1:]
                        logger.info("----------------------------------------------------------------")

                        # Run function and counter check with ML result
                        functionResult = 0
                        emailConFormat = mainFunctions.content_formatting(email.body)
                        emailValidResult = mainFunctions.email_valid(email.from_addr)
                        emailValidResultList.append(emailValidResult)
                        attachmentResult = mainFunctions.attachment_check(emailAttach)
                        attachmentResultList.append(attachmentResult)
                        linkResult = mainFunctions.check_link(email.body)
                        linkResultList.append(linkResult)
                        print("Email valid: ", emailValidResult)
                        print("attachment check: ", attachmentResult)
                        print("Link Check: ", linkResult)

                        # compile result
                        if emailValidResult:
                            functionResult += 25
                        if attachmentResult:
                            functionResult += 25
                        if linkResult:
                            functionResult += 50

                        if functionResult >= 50:
                            function_result = 'Phishing'
                        else:
                            function_result = 'Non-Phishing'

                        result_list.append(ml_result)
                        
                        # insert row in excel
                        ws.insert_rows(2)

                        # read from excel file
                        excelColumn = 'A'
                        excelPosition = excelColumn + str(excelRow)
                        subject_list.append(ws[excelPosition].value)
                        print(excelPosition)

                        excelColumn = chr(ord(excelColumn) + 1)
                        excelPosition = excelColumn + str(excelRow)
                        email_address_list.append(ws[excelPosition].value)
                        print(excelPosition)

                        excelColumn = chr(ord(excelColumn) + 1)
                        excelPosition = excelColumn + str(excelRow)
                        body_list.append(ws[excelPosition].value)
                        print(excelPosition)

                        excelColumn = chr(ord(excelColumn) + 1)
                        excelPosition = excelColumn + str(excelRow)
                        print(excelPosition)

                        excelColumn = chr(ord(excelColumn) + 1)
                        excelPosition = excelColumn + str(excelRow)
                        result_list.append(ws[excelPosition].value)
                        print(excelPosition)

                        excelColumn = chr(ord(excelColumn) + 1)
                        excelPosition = excelColumn + str(excelRow)
                        print(excelPosition)

                        excelColumn = chr(ord(excelColumn) + 1)
                        excelPosition = excelColumn + str(excelRow)
                        print(excelPosition)

                        excelColumn = chr(ord(excelColumn) + 1)
                        excelPosition = excelColumn + str(excelRow)
                        print(excelPosition)

                        excelColumn = chr(ord(excelColumn) + 1)
                        excelPosition = excelColumn + str(excelRow)
                        percentage_list.append(ws[excelPosition].value)
                        print(excelPosition)

                        excelColumn = chr(ord(excelColumn) + 1)
                        excelPosition = excelColumn + str(excelRow)
                        emailValidResultList.append(ws[excelPosition].value)
                        print(excelPosition)

                        excelColumn = chr(ord(excelColumn) + 1)
                        excelPosition = excelColumn + str(excelRow)
                        attachmentResultList.append(ws[excelPosition].value)
                        print(excelPosition)

                        excelColumn = chr(ord(excelColumn) + 1)
                        excelPosition = excelColumn + str(excelRow)
                        linkResultList.append(ws[excelPosition].value)
                        print(excelPosition)

                        excelColumn = chr(ord(excelColumn) + 1)
                        excelPosition = excelColumn + str(excelRow)
                        htmlExistsList.append(ws[excelPosition].value)
                        print(excelPosition)

                        excelColumn = chr(ord(excelColumn) + 1)
                        excelPosition = excelColumn + str(excelRow)
                        domainCountList.append(ws[excelPosition].value)
                        print(excelPosition)

                        excelColumn = chr(ord(excelColumn) + 1)
                        excelPosition = excelColumn + str(excelRow)
                        dotCountList.append(ws[excelPosition].value)
                        print(excelPosition)

                        excelColumn = chr(ord(excelColumn) + 1)
                        excelPosition = excelColumn + str(excelRow)
                        accountExistsList.append(ws[excelPosition].value)
                        print(excelPosition)

                        excelColumn = chr(ord(excelColumn) + 1)
                        excelPosition = excelColumn + str(excelRow)
                        paypalExistsList.append(ws[excelPosition].value)
                        print(excelPosition)

                        excelColumn = chr(ord(excelColumn) + 1)
                        excelPosition = excelColumn + str(excelRow)
                        loginExistsList.append(ws[excelPosition].value)
                        print(excelPosition)

                        excelColumn = chr(ord(excelColumn) + 1)
                        excelPosition = excelColumn + str(excelRow)
                        bankExistsList.append(ws[excelPosition].value)
                        print(excelPosition)

                        excelRow += 1
                        wb.save("logs.xlsx")'''

            # wb.save("logs.xlsx")
            # load into list for web
            wb = load_workbook("logs.xlsx")
            wsNew = wb[username]
            print(wsNew["A2"].value)
            print(excelRow)
            excelRow = 2
            for x in range(1, wsNew.max_row):
                # read from excel file
                excelColumn = 'A'
                excelPosition = excelColumn + str(excelRow)
                subject_list.append(wsNew[excelPosition].value)
                print(excelPosition)

                excelColumn = chr(ord(excelColumn) + 1)
                excelPosition = excelColumn + str(excelRow)
                email_address_list.append(wsNew[excelPosition].value)
                print(excelPosition)

                excelColumn = chr(ord(excelColumn) + 1)
                excelPosition = excelColumn + str(excelRow)
                body_list.append(wsNew[excelPosition].value)
                print(excelPosition)

                excelColumn = chr(ord(excelColumn) + 1)
                excelPosition = excelColumn + str(excelRow)
                print(excelPosition)

                excelColumn = chr(ord(excelColumn) + 1)
                excelPosition = excelColumn + str(excelRow)
                result_list.append(wsNew[excelPosition].value)
                print(excelPosition)

                excelColumn = chr(ord(excelColumn) + 1)
                excelPosition = excelColumn + str(excelRow)
                print(excelPosition)

                excelColumn = chr(ord(excelColumn) + 1)
                excelPosition = excelColumn + str(excelRow)
                print(excelPosition)

                excelColumn = chr(ord(excelColumn) + 1)
                excelPosition = excelColumn + str(excelRow)
                print(excelPosition)

                excelColumn = chr(ord(excelColumn) + 1)
                excelPosition = excelColumn + str(excelRow)
                percentage_list.append(wsNew[excelPosition].value)
                print(excelPosition)

                excelColumn = chr(ord(excelColumn) + 1)
                excelPosition = excelColumn + str(excelRow)
                emailValidResultList.append(ws[excelPosition].value)
                print(excelPosition)

                excelColumn = chr(ord(excelColumn) + 1)
                excelPosition = excelColumn + str(excelRow)
                attachmentResultList.append(ws[excelPosition].value)
                print(excelPosition)

                excelColumn = chr(ord(excelColumn) + 1)
                excelPosition = excelColumn + str(excelRow)
                linkResultList.append(ws[excelPosition].value)
                print(excelPosition)

                excelColumn = chr(ord(excelColumn) + 1)
                excelPosition = excelColumn + str(excelRow)
                htmlExistsList.append(ws[excelPosition].value)
                print(excelPosition)

                excelColumn = chr(ord(excelColumn) + 1)
                excelPosition = excelColumn + str(excelRow)
                domainCountList.append(ws[excelPosition].value)
                print(excelPosition)

                excelColumn = chr(ord(excelColumn) + 1)
                excelPosition = excelColumn + str(excelRow)
                dotCountList.append(ws[excelPosition].value)
                print(excelPosition)

                excelColumn = chr(ord(excelColumn) + 1)
                excelPosition = excelColumn + str(excelRow)
                accountExistsList.append(ws[excelPosition].value)
                print(excelPosition)

                excelColumn = chr(ord(excelColumn) + 1)
                excelPosition = excelColumn + str(excelRow)
                paypalExistsList.append(ws[excelPosition].value)
                print(excelPosition)

                excelColumn = chr(ord(excelColumn) + 1)
                excelPosition = excelColumn + str(excelRow)
                loginExistsList.append(ws[excelPosition].value)
                print(excelPosition)

                excelColumn = chr(ord(excelColumn) + 1)
                excelPosition = excelColumn + str(excelRow)
                bankExistsList.append(ws[excelPosition].value)
                print(excelPosition)

                excelRow += 1

        else:  # create new sheet for new email
            print(username + ' does not exists')
            wb.create_sheet(username)
            print(username + ' created')
            ws = wb[username]
            excelRow = 2
            # server = e.connect(imap_url, username, password)
            # inbox = server.listids()
            for x in range(0, len(inbox)):  # change 10 to len(inbox) to get 100 mails
                email = server.mail(server.listids()[x])
                # store email subject, body in list
                email_address_list.append(email.from_addr)
                subject_list.append(email.title)
                body = mainFunctions.content_formatting(email.body)
                body_list.append(body)


                # ML
                string = body
                string = cleaning(string)
                string = stem(string)
                string = remove_stopwords(string)

                # dataFrame with only the email body that has been cleaned for use with Naive Bayes only
                n_df = pd.DataFrame({'text': string}, index=[0])

                # dataFrame for MLP and RandomForest model (Data not cleaned on purpose)
                df = pd.DataFrame({'text': body}, index=[0])

                # dataFrame with columns extracted from email
                mlp_df = pd.DataFrame(columns=['HtmlExists', 'DomainCount', 'DotCount',
                                               'AccountExists', 'PaypalExists', 'LoginExists', 'BankExists'])

                mlp_df['HtmlExists'] = df['text'].apply(html_exists)
                mlp_df['DomainCount'] = df['text'].apply(count_domain)
                mlp_df['DotCount'] = df['text'].apply(count_dots)
                mlp_df['AccountExists'] = df['text'].apply(account_exists)
                mlp_df['PaypalExists'] = df['text'].apply(paypal_exists)
                mlp_df['LoginExists'] = df['text'].apply(login_exists)
                mlp_df['BankExists'] = df['text'].apply(bank_exists)
                print(mlp_df)

                # save columns to variable
                htmlExists = mlp_df.iloc[0]['HtmlExists']
                domainCount = mlp_df.iloc[0]['DomainCount']
                dotCount = mlp_df.iloc[0]['DotCount']
                accountExists = mlp_df.iloc[0]['AccountExists']
                paypalExists = mlp_df.iloc[0]['PaypalExists']
                loginExists = mlp_df.iloc[0]['LoginExists']
                bankExists = mlp_df.iloc[0]['BankExists']

                # append to list
                htmlExistsList.append(htmlExists)
                domainCountList.append(domainCount)
                dotCountList.append(dotCount)
                accountExistsList.append(accountExists)
                paypalExistsList.append(paypalExists)
                loginExistsList.append(loginExists)
                bankExistsList.append(bankExists)

                if model == 0:  # naive bayes
                    vectorizer = load(r'naivebayesVectorizer.joblib')  # load vectorizer
                    nbclf = load(r'naivebayes.joblib')  # load the naivebayes ml model
                    x_matrix = n_df['text']
                    x_matrix = vectorizer.transform(n_df['text'])
                    my_prediction = nbclf.predict(x_matrix)
                    percentage = nbclf.predict_proba(x_matrix)
                    print(my_prediction, "MY PREDICTION")

                if model == 1:  # MLP
                    nbclf = load(r'mlp.joblib')
                    my_prediction = nbclf.predict(mlp_df)
                    percentage = nbclf.predict_proba(mlp_df)
                    print(my_prediction, "MY PREDICTION")

                if model == 2:  # Random Forest
                    nbclf = load(r'randomforest.joblib')
                    my_prediction = nbclf.predict(mlp_df)
                    percentage = nbclf.predict_proba(mlp_df)
                    print(my_prediction, "MY PREDICTION")

                np.set_printoptions(formatter={'float_kind': '{:f}'.format})

                if my_prediction == 1:
                    ml_result = 'Phishing'
                    percentage = format(percentage[0][1], '.12f')  # to 12decimal place
                    percentage = float(percentage) * 100  # convert to percent
                    percentage = str(percentage) + '%'
                    percentage_list.append(percentage)
                elif my_prediction == 0:
                    ml_result = 'Non-Phishing'
                    percentage = format(percentage[0][0], '.12f')  # to 12decimal place
                    percentage = float(percentage) * 100  # convert to percent
                    percentage = str(percentage) + '%'
                    percentage_list.append(percentage)

                # logger.info("Email attachment: ")
                # logger.info(email.attachments)
                emailAttachment = email.attachments
                if not emailAttachment:
                    emailAttach = "-"
                else:
                    attachment = emailAttachment[0]
                    attachment = str(attachment)
                    attach = attachment.split(',')
                    emailAttach = str(attach[0])
                    emailAttach = emailAttach[1:]
                logger.info("----------------------------------------------------------------")

                # Run function and counter check with ML result
                functionResult = 0
                emailConFormat = mainFunctions.content_formatting(email.body)
                emailValidResult = mainFunctions.email_valid(email.from_addr)
                emailValidResultList.append(emailValidResult)
                attachmentResult = mainFunctions.attachment_check(emailAttach)
                attachmentResultList.append(attachmentResult)
                linkResult = mainFunctions.check_link(email.body)
                linkResultList.append(linkResult)
                print("Email valid: ", emailValidResult)
                print("attachment check: ", attachmentResult)
                print("Link Check: ", linkResult)

                # compile result
                if emailValidResult:
                    functionResult += 25
                if attachmentResult:
                    functionResult += 25
                if linkResult:
                    functionResult += 50

                if functionResult >= 50:
                    function_result = 'Phishing'
                else:
                    function_result = 'Non-Phishing'

                result_list.append(ml_result)

                # Check if email address is in user blacklist or whitelist
                listing = "-"
                if email.from_addr in userBlacklist:
                    listing = "blacklist"
                if email.from_addr in userWhitelist:
                    listing = "blacklist"

                # excel file titles
                excelColumn1 = 'A1'
                excelPosition1 = excelColumn1
                ws[excelPosition1] = "Email Subject"
                # ws.write(excelPosition1, "Email Subject")
                print(excelPosition1)

                excelColumn2 = 'B1'
                excelPosition2 = excelColumn2
                ws[excelPosition2] = "Name and Email Address"
                # ws.write(excelPosition2, "Name and Email Address")
                print(excelPosition2)

                excelColumn3 = 'C1'
                excelPosition3 = excelColumn3
                ws[excelPosition3] = "Email Content"
                # ws.write(excelPosition3, "Email Content")
                print(excelPosition3)

                excelColumn4 = 'D1'
                excelPosition4 = excelColumn4
                ws[excelPosition4] = "Listing"
                # ws.write(excelPosition4, "Listing")
                print(excelPosition4)

                excelColumn5 = 'E1'
                excelPosition5 = excelColumn5
                ws[excelPosition5] = "Classification"
                # ws.write(excelPosition5, "Classification")
                print(excelPosition5)

                excelColumn6 = 'F1'
                excelPosition6 = excelColumn6
                ws[excelPosition6] = "Attachment"
                # ws.write(excelPosition6, "Attachment")
                print(excelPosition6)

                excelColumn7 = 'G1'
                excelPosition7 = excelColumn7
                ws[excelPosition7] = "Ml Result"
                # ws.write(excelPosition6, "ML Result")
                print(excelPosition7)

                excelColumn8 = 'H1'
                excelPosition8 = excelColumn8
                ws[excelPosition8] = "Function Result"
                # ws.write(excelPosition8, "Function Result")
                print(excelPosition8)

                excelColumn9 = 'I1'
                excelPosition9 = excelColumn9
                ws[excelPosition9] = "Percentage"
                # ws.write(excelPosition9, "Percentage")
                print(excelPosition9)

                excelColumn10 = 'J1'
                excelPosition10 = excelColumn10
                ws[excelPosition10] = "Email Valid"
                # ws.write(excelPosition10, "Email Valid")
                print(excelPosition10)

                excelColumn11 = 'K1'
                excelPosition11 = excelColumn11
                ws[excelPosition11] = "Attachment Result"
                # ws.write(excelPosition11, "Attachment Result")
                print(excelPosition11)

                excelColumn12 = 'L1'
                excelPosition12 = excelColumn12
                ws[excelPosition12] = "Link Result"
                # ws.write(excelPosition12, "Link Result")
                print(excelPosition12)

                excelColumn13 = 'M1'
                excelPosition13 = excelColumn13
                ws[excelPosition13] = "HTML Exists"
                # ws.write(excelPosition13, "HTML Exists")
                print(excelPosition13)

                excelColumn14 = 'N1'
                excelPosition14 = excelColumn14
                ws[excelPosition14] = "Domain Count"
                # ws.write(excelPosition14, "Domain Count")
                print(excelPosition14)

                excelColumn15 = 'O1'
                excelPosition15 = excelColumn15
                ws[excelPosition15] = "Dot Count"
                # ws.write(excelPosition15, "Dot Count")
                print(excelPosition15)

                excelColumn16 = 'P1'
                excelPosition16 = excelColumn16
                ws[excelPosition16] = "Account Exists"
                # ws.write(excelPosition16, "Account Exists")
                print(excelPosition16)

                excelColumn17 = 'Q1'
                excelPosition17 = excelColumn17
                ws[excelPosition17] = "Paypal Exist"
                # ws.write(excelPosition17, "Paypal exists")
                print(excelPosition17)

                excelColumn18 = 'R1'
                excelPosition18 = excelColumn18
                ws[excelPosition18] = "Login Exists"
                # ws.write(excelPosition18, "Login exists")
                print(excelPosition18)

                excelColumn19 = 'S1'
                excelPosition19 = excelColumn19
                ws[excelPosition19] = "Bank Exists"
                # ws.write(excelPosition19, "Bank Exists")
                print(excelPosition19)

                # excel file
                excelColumn = 'A'
                excelPosition = excelColumn + str(excelRow)
                ws[excelPosition] = email.title
                # ws.write(excelPosition, email.title)
                print(excelPosition)

                excelColumn = chr(ord(excelColumn) + 1)
                excelPosition = excelColumn + str(excelRow)
                ws[excelPosition] = email.from_addr
                # ws.write(excelPosition, email.from_addr)
                print(excelPosition)

                excelColumn = chr(ord(excelColumn) + 1)
                excelPosition = excelColumn + str(excelRow)
                ws[excelPosition] = emailConFormat
                # ws.write(excelPosition, emailConFormat)
                print(excelPosition)

                excelColumn = chr(ord(excelColumn) + 1)
                excelPosition = excelColumn + str(excelRow)
                ws[excelPosition] = listing
                # ws.write(excelPosition, "-")
                print(excelPosition)

                excelColumn = chr(ord(excelColumn) + 1)
                excelPosition = excelColumn + str(excelRow)
                ws[excelPosition] = ml_result
                # ws.write(excelPosition, result)
                print(excelPosition)

                excelColumn = chr(ord(excelColumn) + 1)
                excelPosition = excelColumn + str(excelRow)
                ws[excelPosition] = emailAttach
                # ws.write(excelPosition, emailAttach)
                print(excelPosition)

                excelColumn = chr(ord(excelColumn) + 1)
                excelPosition = excelColumn + str(excelRow)
                ws[excelPosition] = ml_result
                # ws.write(excelPosition, ml_result)
                print(excelPosition)

                excelColumn = chr(ord(excelColumn) + 1)
                excelPosition = excelColumn + str(excelRow)
                ws[excelPosition] = function_result
                # ws.write(excelPosition, function_result)
                print(excelPosition)

                excelColumn = chr(ord(excelColumn) + 1)
                excelPosition = excelColumn + str(excelRow)
                ws[excelPosition] = percentage
                # ws.write(excelPosition, percentage)
                print(excelPosition)

                excelColumn = chr(ord(excelColumn) + 1)
                excelPosition = excelColumn + str(excelRow)
                ws[excelPosition] = emailValidResult
                print(excelPosition)

                excelColumn = chr(ord(excelColumn) + 1)
                excelPosition = excelColumn + str(excelRow)
                ws[excelPosition] = attachmentResult
                print(excelPosition)

                excelColumn = chr(ord(excelColumn) + 1)
                excelPosition = excelColumn + str(excelRow)
                ws[excelPosition] = linkResult
                print(excelPosition)

                excelColumn = chr(ord(excelColumn) + 1)
                excelPosition = excelColumn + str(excelRow)
                ws[excelPosition] = htmlExists
                print(excelPosition)

                excelColumn = chr(ord(excelColumn) + 1)
                excelPosition = excelColumn + str(excelRow)
                ws[excelPosition] = domainCount
                print(excelPosition)

                excelColumn = chr(ord(excelColumn) + 1)
                excelPosition = excelColumn + str(excelRow)
                ws[excelPosition] = dotCount
                print(excelPosition)

                excelColumn = chr(ord(excelColumn) + 1)
                excelPosition = excelColumn + str(excelRow)
                ws[excelPosition] = accountExists
                print(excelPosition)

                excelColumn = chr(ord(excelColumn) + 1)
                excelPosition = excelColumn + str(excelRow)
                ws[excelPosition] = paypalExists
                print(excelPosition)

                excelColumn = chr(ord(excelColumn) + 1)
                excelPosition = excelColumn + str(excelRow)
                ws[excelPosition] = loginExists
                print(excelPosition)

                excelColumn = chr(ord(excelColumn) + 1)
                excelPosition = excelColumn + str(excelRow)
                ws[excelPosition] = bankExists
                print(excelPosition)
            # ws.close()
            wb.save("logs.xlsx")

    else:
        # excel file
        excel = xlsxwriter.Workbook('logs.xlsx')
        bold = excel.add_format({'bold': True})
        worksheet = excel.add_worksheet(username)
        excelRow = 2

        # server = e.connect(imap_url, username, password)
        # inbox = server.listup()
        # inbox = server.listids()
        email = server.mail(server.listids()[0])

        for x in range(0, len(inbox)):  # change 10 to len(inbox) to get 100 mails
            email = server.mail(server.listids()[x])

            # store email subject, body in list
            email_address_list.append(email.from_addr)
            subject_list.append(email.title)
            body = mainFunctions.content_formatting(email.body)
            body_list.append(body)

            # ML
            string = body
            string = cleaning(string)
            string = stem(string)
            string = remove_stopwords(string)

            # dataFrame with only the email body
            n_df = pd.DataFrame({'text': string}, index=[0])

            # dataFrame for MLP and RandomForest model (Data not cleaned on purpose)
            df = pd.DataFrame({'text': body}, index=[0])

            # dataFrame with columns extracted from email
            mlp_df = pd.DataFrame(columns=['HtmlExists', 'DomainCount', 'DotCount',
                                           'AccountExists', 'PaypalExists', 'LoginExists', 'BankExists'])

            # perform function and create columns
            mlp_df['HtmlExists'] = df['text'].apply(html_exists)
            mlp_df['DomainCount'] = df['text'].apply(count_domain)
            mlp_df['DotCount'] = df['text'].apply(count_dots)
            mlp_df['AccountExists'] = df['text'].apply(account_exists)
            mlp_df['PaypalExists'] = df['text'].apply(paypal_exists)
            mlp_df['LoginExists'] = df['text'].apply(login_exists)
            mlp_df['BankExists'] = df['text'].apply(bank_exists)
            print(mlp_df)

            # save value to variable
            htmlExists = mlp_df.iloc[0]['HtmlExists']
            domainCount = mlp_df.iloc[0]['DomainCount']
            dotCount = mlp_df.iloc[0]['DotCount']
            accountExists = mlp_df.iloc[0]['AccountExists']
            paypalExists = mlp_df.iloc[0]['PaypalExists']
            loginExists = mlp_df.iloc[0]['LoginExists']
            bankExists = mlp_df.iloc[0]['BankExists']

            # append to list
            htmlExistsList.append(htmlExists)
            domainCountList.append(domainCount)
            dotCountList.append(dotCount)
            accountExistsList.append(accountExists)
            paypalExistsList.append(paypalExists)
            loginExistsList.append(loginExists)
            bankExistsList.append(bankExists)

            if model == 0:  # naive bayes
                vectorizer = load(r'naivebayesVectorizer.joblib')  # load vectorizer
                nbclf = load(r'naivebayes.joblib')  # load the naivebayes ml model
                x_matrix = n_df['text']
                x_matrix = vectorizer.transform(n_df['text'])
                my_prediction = nbclf.predict(x_matrix)
                percentage = nbclf.predict_proba(x_matrix)
                print(my_prediction, "MY PREDICTION")

            if model == 1:  # MLP
                nbclf = load(r'mlp.joblib')
                my_prediction = nbclf.predict(mlp_df)
                percentage = nbclf.predict_proba(mlp_df)
                print(my_prediction, "MY PREDICTION")

            if model == 2:  # Random Forest
                nbclf = load(r'randomforest.joblib')
                my_prediction = nbclf.predict(mlp_df)
                percentage = nbclf.predict_proba(mlp_df)
                print(my_prediction, "MY PREDICTION")

            # percentage = np.array(percentage)
            # percentage = ['{:f}'.format(item) for item in percentage]
            np.set_printoptions(formatter={'float_kind': '{:f}'.format})

            if my_prediction == 1:
                ml_result = 'Phishing'
                percentage = format(percentage[0][1], '.12f')  # to 12decimal place
                percentage = float(percentage) * 100  # convert to percent
                percentage = str(percentage) + '%'
                percentage_list.append(percentage)
            elif my_prediction == 0:
                ml_result = 'Non-Phishing'
                percentage = format(percentage[0][0], '.12f')  # to 12decimal place
                percentage = float(percentage) * 100  # convert to percent
                percentage = str(percentage) + '%'
                percentage_list.append(percentage)

            # logger.info("Email attachment: ")
            # logger.info(email.attachments)
            emailAttachment = email.attachments
            if not emailAttachment:
                emailAttach = "-"
            else:
                attachment = emailAttachment[0]
                attachment = str(attachment)
                attach = attachment.split(',')
                emailAttach = str(attach[0])
                emailAttach = emailAttach[1:]
            logger.info("----------------------------------------------------------------")

            # Run function and counter check with ML result
            functionResult = 100
            emailConFormat = mainFunctions.content_formatting(email.body)
            emailValidResult = mainFunctions.email_valid(email.from_addr)
            emailValidResultList.append(emailValidResult)
            attachmentResult = mainFunctions.attachment_check(emailAttach)
            attachmentResultList.append(attachmentResult)
            linkResult = mainFunctions.check_link(email.body)
            linkResultList.append(linkResult)
            print("Email valid: ", emailValidResult)
            print("attachment check: ", attachmentResult)
            print("Link Check: ", linkResult)

            # compile result
            if emailValidResult:
                functionResult += 25
            if attachmentResult:
                functionResult += 25
            if linkResult:
                functionResult += 50

            if functionResult >= 50:
                function_result = 'Phishing'
            else:
                function_result = 'Non-Phishing'

            result_list.append(ml_result)

            # Check if email address is in user blacklist or whitelist
            listing = "-"
            if email.from_addr in userBlacklist:
                listing = "blacklist"
            if email.from_addr in userWhitelist:
                listing = "blacklist"

            # excel file titles
            excelColumn1 = 'A1'
            excelPosition1 = excelColumn1
            worksheet.write(excelPosition1, "Email Subject", bold)
            print(excelPosition1)

            excelColumn2 = 'B1'
            excelPosition2 = excelColumn2
            worksheet.write(excelPosition2, "Name and Email Address", bold)
            print(excelPosition2)

            excelColumn3 = 'C1'
            excelPosition3 = excelColumn3
            worksheet.write(excelPosition3, "Email Content", bold)
            print(excelPosition3)

            excelColumn4 = 'D1'
            excelPosition4 = excelColumn4
            worksheet.write(excelPosition4, "Listing", bold)
            print(excelPosition4)

            excelColumn5 = 'E1'
            excelPosition5 = excelColumn5
            worksheet.write(excelPosition5, "Classification", bold)
            print(excelPosition5)

            excelColumn6 = 'F1'
            excelPosition6 = excelColumn6
            worksheet.write(excelPosition6, "Attachment", bold)
            print(excelPosition6)

            excelColumn7 = 'G1'
            excelPosition7 = excelColumn7
            worksheet.write(excelPosition7, "ML Result", bold)
            print(excelPosition7)

            excelColumn8 = 'H1'
            excelPosition8 = excelColumn8
            worksheet.write(excelPosition8, "Function Result", bold)
            print(excelPosition8)

            excelColumn9 = 'I1'
            excelPosition9 = excelColumn9
            worksheet.write(excelPosition9, "Percentage", bold)
            print(excelPosition9)

            excelColumn10 = 'J1'
            excelPosition10 = excelColumn10
            worksheet.write(excelPosition10, "Email Valid", bold)
            print(excelPosition10)

            excelColumn11 = 'K1'
            excelPosition11 = excelColumn11
            worksheet.write(excelPosition11, "Attachment Result", bold)
            print(excelPosition11)

            excelColumn12 = 'L1'
            excelPosition12 = excelColumn12
            worksheet.write(excelPosition12, "Link Result", bold)
            print(excelPosition12)

            excelColumn13 = 'M1'
            excelPosition13 = excelColumn13
            worksheet.write(excelPosition13, "HTML Exists", bold)
            print(excelPosition13)

            excelColumn14 = 'N1'
            excelPosition14 = excelColumn14
            worksheet.write(excelPosition14, "Domain Count", bold)
            print(excelPosition14)

            excelColumn15 = 'O1'
            excelPosition15 = excelColumn15
            worksheet.write(excelPosition15, "Dot Count", bold)
            print(excelPosition15)

            excelColumn16 = 'P1'
            excelPosition16 = excelColumn16
            worksheet.write(excelPosition16, "Account Exists", bold)
            print(excelPosition16)

            excelColumn17 = 'Q1'
            excelPosition17 = excelColumn17
            worksheet.write(excelPosition17, "Paypal Exists", bold)
            print(excelPosition17)

            excelColumn18 = 'R1'
            excelPosition18 = excelColumn18
            worksheet.write(excelPosition18, "Login Exists", bold)
            print(excelPosition18)

            excelColumn19 = 'S1'
            excelPosition19 = excelColumn19
            worksheet.write(excelPosition19, "Bank Exists", bold)
            print(excelPosition19)

            # excel file
            excelColumn = 'A'
            excelPosition = excelColumn + str(excelRow)
            worksheet.write(excelPosition, email.title)
            print(excelPosition)

            excelColumn = chr(ord(excelColumn) + 1)
            excelPosition = excelColumn + str(excelRow)
            worksheet.write(excelPosition, email.from_addr)
            print(excelPosition)

            excelColumn = chr(ord(excelColumn) + 1)
            excelPosition = excelColumn + str(excelRow)
            worksheet.write(excelPosition, emailConFormat)
            print(excelPosition)

            excelColumn = chr(ord(excelColumn) + 1)
            excelPosition = excelColumn + str(excelRow)
            worksheet.write(excelPosition, listing)
            print(excelPosition)

            excelColumn = chr(ord(excelColumn) + 1)
            excelPosition = excelColumn + str(excelRow)
            worksheet.write(excelPosition, ml_result)
            print(excelPosition)

            excelColumn = chr(ord(excelColumn) + 1)
            excelPosition = excelColumn + str(excelRow)
            worksheet.write(excelPosition, emailAttach)
            print(excelPosition)

            excelColumn = chr(ord(excelColumn) + 1)
            excelPosition = excelColumn + str(excelRow)
            worksheet.write(excelPosition, ml_result)
            print(excelPosition)

            excelColumn = chr(ord(excelColumn) + 1)
            excelPosition = excelColumn + str(excelRow)
            worksheet.write(excelPosition, function_result)
            print(excelPosition)

            excelColumn = chr(ord(excelColumn) + 1)
            excelPosition = excelColumn + str(excelRow)
            worksheet.write(excelPosition, percentage)
            print(excelPosition)

            excelColumn = chr(ord(excelColumn) + 1)
            excelPosition = excelColumn + str(excelRow)
            worksheet.write(excelPosition, emailValidResult)
            print(excelPosition)

            excelColumn = chr(ord(excelColumn) + 1)
            excelPosition = excelColumn + str(excelRow)
            worksheet.write(excelPosition, attachmentResult)
            print(excelPosition)

            excelColumn = chr(ord(excelColumn) + 1)
            excelPosition = excelColumn + str(excelRow)
            worksheet.write(excelPosition, linkResult)
            print(excelPosition)

            excelColumn = chr(ord(excelColumn) + 1)
            excelPosition = excelColumn + str(excelRow)
            worksheet.write(excelPosition, htmlExists)
            print(excelPosition)

            excelColumn = chr(ord(excelColumn) + 1)
            excelPosition = excelColumn + str(excelRow)
            worksheet.write(excelPosition, domainCount)
            print(excelPosition)

            excelColumn = chr(ord(excelColumn) + 1)
            excelPosition = excelColumn + str(excelRow)
            worksheet.write(excelPosition, dotCount)
            print(excelPosition)

            excelColumn = chr(ord(excelColumn) + 1)
            excelPosition = excelColumn + str(excelRow)
            worksheet.write(excelPosition, accountExists)
            print(excelPosition)

            excelColumn = chr(ord(excelColumn) + 1)
            excelPosition = excelColumn + str(excelRow)
            worksheet.write(excelPosition, paypalExists)
            print(excelPosition)

            excelColumn = chr(ord(excelColumn) + 1)
            excelPosition = excelColumn + str(excelRow)
            worksheet.write(excelPosition, loginExists)
            print(excelPosition)

            excelColumn = chr(ord(excelColumn) + 1)
            excelPosition = excelColumn + str(excelRow)
            worksheet.write(excelPosition, bankExists)
            print(excelPosition)

            excelRow += 1
        excel.close()


    server = smtplib.SMTP('smtp.gmail.com', 587)  # smtp settings, change accordingly.
    server.ehlo()
    server.starttls()  # secure connection
    bodyList = body_list[0].replace("_x000D_", "")
    specificList = "empty"
    badge = 2
    for x in userBlacklist:
        if x in email_address_list[0]:
            specificList = "Blacklist"
            badge = 0  # red
    for z in userWhitelist:
        if z in email_address_list[0]:
            specificList = "Whitelist"
            badge = 1

    return render_template("inbox.html", len=len(subject_list), subject=subject_list,
                           address=email_address_list, body=bodyList, userList=specificList, badge=badge,
                           result_list=result_list, percentage_list=percentage_list, model_string=model_string)


@app.route('/inbox/<num>')
def showEmail(num):
    getresult = result_list[int(num)]
    getpercentage = percentage_list[int(num)]
    specific_subject = subject_list[int(num)]
    specific_address = email_address_list[int(num)]
    specific_body = body_list[int(num)]
    print(specific_subject)
    specific_body = specific_body.replace("_x000D_", "")
    specificList = "empty"
    badge = 2
    for x in userBlacklist:
        if x in specific_address:
            specificList = "Blacklist"
            badge = 0  # red
    for z in userWhitelist:
        if z in specific_address:
            specificList = "Whitelist"
            badge = 1

    return render_template("inbox1.html", len=len(subject_list), subject=subject_list,
                           address=email_address_list, body=body_list, num=num,
                           result=getresult, percentage=getpercentage, userList=specificList, badge=badge,
                           specific_body=specific_body, specific_subject=specific_subject,
                           specific_address=specific_address, model_string=model_string
                           )

@app.route('/inbox/analysis')
def analysis():
    emailValid = emailValidResultList[0]
    if emailValid == True:
        emailValid = 'Yes'
    if emailValid == False:
        emailValid = 'No'

    attachment = attachmentResultList[0]
    if attachment == True:
        attachment = "No"
    if attachment == False:
        attachment = "Yes"

    link = linkResultList[0]
    if link == True:
        link = "No"
    if link == False:
        link = "Yes"

    htmlExist = htmlExistsList[0]
    domainCount = domainCountList[0]
    dotCount = dotCountList[0]
    accountExist = accountExistsList[0]
    paypalExist = paypalExistsList[0]
    loginExist = loginExistsList[0]
    bankExist = bankExistsList[0]

    result = result_list[0]
    percentage = percentage_list[0]

    #mlp_df.to_csv('analysis.csv')
    return render_template('analysis.html', emailValid=emailValid, attachment=attachment, link=link, htmlExist=htmlExist,
                           domainCount=domainCount, dotCount=dotCount, accountExist=accountExist, paypalExist=paypalExist,
                           loginExist=loginExist, bankExist=bankExist, result=result, percentage=percentage)


@app.route('/inbox/<int:num>/analysis')
def analysis1(num):
    emailValid = emailValidResultList[num]
    attachment = attachmentResultList[num]
    link = linkResultList[num]

    htmlExist = htmlExistsList[num]
    domainCount = domainCountList[num]
    dotCount = dotCountList[num]
    accountExist = accountExistsList[num]
    paypalExist = paypalExistsList[num]
    loginExist = loginExistsList[num]
    bankExist = bankExistsList[num]

    result = result_list[num]
    percentage = percentage_list[num]

    return render_template('analysis1.html', emailValid=emailValid, attachment=attachment, link=link,
                           htmlExist=htmlExist,
                           domainCount=domainCount, dotCount=dotCount, accountExist=accountExist,
                           paypalExist=paypalExist,
                           loginExist=loginExist, bankExist=bankExist, result=result, percentage=percentage)


@app.route('/inbox/blacklist')
def blacklist():
    global blacklist
    blacklist = []
    wb = load_workbook('blacklist.xlsx')
    if username in wb.sheetnames:
        sheet = wb[username]
    else:
        # wb.create_sheet(username)
        wb.copy_worksheet(wb["blacklist"]).title = username
        wb.save('blacklist.xlsx')
        sheet = wb[username]

    row_count = sheet.max_row
    # nested list of email address and status
    # nested list will look like this in the end [[email,status], [email1,status1], [email2,status2]]
    for i in range(1, row_count + 1):
        for k in range(1, 3):
            if k == 1:
                emailadd = sheet.cell(row=i, column=k).value
                list1 = []
                list1.append(emailadd)
                userBlacklist.append(emailadd)
            if k == 2:
                status = sheet.cell(row=i, column=k).value
                list1.append(status)
                blacklist.append(list1)


    return render_template("blacklist.html", list=blacklist)


@app.route('/inbox/blacklist/new', methods=['GET', 'POST'])
def blacklistnew():
    wb = load_workbook('blacklist.xlsx')
    if username in wb.sheetnames:
        sheet = wb[username]
    else:
        # wb.create_sheet(username)
        wb.copy_worksheet(wb["blacklist"]).title = username
        wb.save('blacklist.xlsx')
        sheet = wb[username]
    # sheet = wb["blacklist"]  # values will be saved to excel sheet"blacklist"
    col2 = 'blacklisted'  # value of 2nd column in excel
    if request.method == 'POST':
        email1 = request.form['email1']  # id='email1' from html form
        if email1.strip() != "":
            col1 = email1.strip()
            sheet.append([col1, col2])

        email2 = request.form['email2']
        if email2.strip() != "":
            col1 = email2.strip()
            sheet.append([col1, col2])

        email3 = request.form['email3']
        if email3.strip() != "":
            col1 = email3.strip()
            sheet.append([col1, col2])

        email4 = request.form['email4']
        if email4.strip() != "":
            col1 = email4.strip()
            sheet.append([col1, col2])
        wb.save('blacklist.xlsx')
        wb.close()
        return redirect('/inbox/blacklist')
    else:
        return render_template("blacklistnew.html")

@app.route('/inbox/blacklist/remove/<email>')
def removeBlacklist(email):
    wb = load_workbook('blacklist.xlsx')
    if username in wb.sheetnames:
        sheet = wb[username]
    else:
        # wb.create_sheet(username)
        wb.copy_worksheet(wb["blacklist"]).title = username
        wb.save('blacklist.xlsx')
        sheet = wb[username]
    # sheet = wb["blacklist"]  # excel sheet"blacklist"
    row_count = sheet.max_row
    k = 1
    print("This is email: ", email)

    # itterate rows in excel sheet
    for i in range(1, row_count + 1):
        emailadd = sheet.cell(row=i, column=k).value
        print(emailadd)
        if emailadd == email:
            row_number = i
            sheet.delete_rows(i, 1)
            wb.save('blacklist.xlsx')
            wb.close()
    return redirect('/inbox/blacklist')

@app.route('/inbox/whitelist')
def whitelist():
    wb = load_workbook('whitelist.xlsx')
    if username in wb.sheetnames:
        sheet = wb[username]
    else:
        # wb.create_sheet(username)
        wb.copy_worksheet(wb["whitelist"]).title = username
        wb.save('whitelist.xlsx')
        sheet = wb[username]

    row_count = sheet.max_row
    list = []  # nested list of email address and status
    # nested list will look like this in the end [[email,status], [email1,status1], [email2,status2]]
    for i in range(1, row_count + 1):
        for k in range(1, 3):
            if k == 1:
                emailadd = sheet.cell(row=i, column=k).value
                list1 = []
                list1.append(emailadd)
                userWhitelist.append(emailadd)
            if k == 2:
                status = sheet.cell(row=i, column=k).value
                list1.append(status)
                list.append(list1)

    return render_template("whitelist.html", list=list)


@app.route('/inbox/whitelist/new', methods=['GET', 'POST'])
def whitelistnew():
    wb = load_workbook('whitelist.xlsx')
    if username in wb.sheetnames:
        sheet = wb[username]
    else:
        # wb.create_sheet(username)
        wb.copy_worksheet(wb["whitelist"]).title = username
        wb.save('whitelist.xlsx')
        sheet = wb[username]

    col2 = 'whitelisted'  # value of 2nd column in excel
    if request.method == 'POST':
        email1 = request.form['email1']  # id='email1' from html form
        if email1.strip() != "":
            col1 = email1.strip()
            sheet.append([col1, col2])

        email2 = request.form['email2']
        if email2.strip() != "":
            col1 = email2.strip()
            sheet.append([col1, col2])

        email3 = request.form['email3']
        if email3.strip() != "":
            col1 = email3.strip()
            sheet.append([col1, col2])

        email4 = request.form['email4']
        if email4.strip() != "":
            col1 = email4.strip()
            sheet.append([col1, col2])
        wb.save('whitelist.xlsx')
        wb.close()
        return redirect('/inbox/whitelist')
    else:
        return render_template("whitelistnew.html")


@app.route('/inbox/whitelist/remove/<email>')
def removeWhitelist(email):
    print("This is email: ", email)
    wb = load_workbook('whitelist.xlsx')
    if username in wb.sheetnames:
        sheet = wb[username]
    else:
        # wb.create_sheet(username)
        wb.copy_worksheet(wb["whitelist"]).title = username
        wb.save('whitelist.xlsx')
        sheet = wb[username]

    row_count = sheet.max_row
    k = 1

    # itterate rows in excel sheet
    for i in range(1, row_count + 1):
        emailadd = sheet.cell(row=i, column=k).value
        print(emailadd)
        if emailadd == email:
            row_number = i
            sheet.delete_rows(i, 1)
            wb.save('whitelist.xlsx')
            wb.close()
    return redirect('/inbox/whitelist')

@app.route('/inbox/quarantine')
def showQuarantine():
    from openpyxl import load_workbook
    wb = load_workbook('logs.xlsx')
    ws = wb[username]
    global percentageList
    global subjectList
    global bodyList
    global emailAddressList
    global resultList

    percentageList = []
    subjectList = []
    bodyList = []
    emailAddressList = []
    resultList = []

    for row in ws.rows:
        if row[4].value == "Phishing":
            subjectList.append(row[0].value)
            emailAddressList.append(row[1].value)
            bodyList.append(row[2].value)
            resultList.append(row[4].value)
            percentageList.append(row[8].value)
    bodylist = bodyList[0].replace("_x000D_", "")
    specificList = "empty"
    badge = 2
    for x in userBlacklist:
        if x in emailAddressList[0]:
            specificList = "Blacklist"
            badge = 0  # red
    for z in userWhitelist:
        if z in emailAddressList[0]:
            specificList = "Whitelist"
            badge = 1

    return render_template("quarantine.html", len=len(subjectList), subject=subjectList,
                           address=emailAddressList, body=bodylist, userList=specificList, badge=badge,
                           result_list=resultList, percentage_list=percentageList, model_string=model_string)

@app.route('/inbox/quarantine/<num>')
def showSuspicious(num):
    getresult = resultList[int(num)]
    getpercentage = percentageList[int(num)]
    specific_subject = subjectList[int(num)]
    specific_address = emailAddressList[int(num)]
    specific_body = bodyList[int(num)]
    specific_body = specific_body.replace("_x000D_", "")
    specificList = "empty"
    badge = 2
    for x in userBlacklist:
        if x in specific_address:
            specificList = "Blacklist"
            badge = 0  # red
    for z in userWhitelist:
        if z in specific_address:
            specificList = "Whitelist"
            badge = 1

    return render_template("quarantine1.html", len=len(subjectList), subject=subjectList,
                           address=emailAddressList, num=num, model_string=model_string,
                           result=getresult, percentage=getpercentage, userList=specificList, badge=badge,
                           specific_body=specific_body, specific_subject=specific_subject,
                           specific_address=specific_address
                           )

@app.route('/model/<int:num>')
def model(num):
    wb = load_workbook('model.xlsx')
    sheet = wb["Sheet1"]  # values will be saved to excel sheet"blacklist"
    sheet['A1'] = num

    wb.save('model.xlsx')
    wb.close()

    return redirect("/inbox")

@app.route("/logout")
def logout():
    server.quit()
    username = None
    password = None
    return redirect("/")


# to run application
if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8080, debug=False)
    #app.run()
